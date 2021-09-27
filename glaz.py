from telethon import TelegramClient, events
import asyncio
from openpyxl import Workbook, load_workbook
from enum import Enum, auto
import json
from loguru import logger
import glob
import sqlite3

# Use your own values from my.telegram.org
API_ID = 1018957
API_HASH = '4b0f0fe8a71dbff9a4422dc1ab43d4f2'



class NotSubed(Exception):
    """У аккаунта нет подписки"""
    pass

class NotExpectedMsg(Exception):
    """Сообщение не того формата, который ожидался"""
    pass

class Archive():
    """Статический класс управления архивом сообщений"""
    @staticmethod
    def get_conn():
        return sqlite3.connect("archive.db")

    @staticmethod
    def create_table():
        conn = Archive.get_conn()
        cursor = conn.cursor()
        cursor.execute("""  
            CREATE TABLE IF NOT EXISTS archive
            (number text UNIQUE, msg text)""")
        conn.commit()

    @staticmethod
    def add_to_archive(number, msg):
        conn = Archive.get_conn()
        cursor = conn.cursor()
        cursor.execute("""  
            INSERT OR IGNORE INTO archive
            VALUES (?, ?)""",(number, msg))
        conn.commit()

    @staticmethod
    def get_from_archive(number):
        conn = Archive.get_conn()
        cursor = conn.cursor()
        cursor.execute(f"""  
            SELECT msg FROM archive WHERE number = '{number}'""")
        result = cursor.fetchall()
        if result:
            return result[0][0]
        else:
            return False
class Parser():
    """Статический класс парсинга информации из сообщения"""
    @staticmethod
    def get_attempts(msg):
        sub_msg = Parser.get_field(msg, "Подписка до")
        if not sub_msg:
            raise NotExpectedMsg("bad info msg format")
        if sub_msg.find("отсутствует") != -1:
            raise NotSubed()
        attempts_msg = Parser.get_field(msg, "Лимит запросов в сутки")
        try:
            spl_msg = attempts_msg.split("/")
            return int(spl_msg[1]) - int(spl_msg[0])
        except ValueError:
            raise NotExpectedMsg("bad info msg format")

    @staticmethod
    def get_data(msg, pattern):
        """Основной метод парсинга. Возвращает список из полей pattern"""
        data = []
        if "Если информация не найдена, закажите" not in msg:
            raise NotExpectedMsg("Bad msg format")
        info = Parser.get_field(msg, "Номер")
        data.append(info)
        for field_type in pattern:
            if field_type == "fio":
                info = Parser.get_field(msg, "ФИО")
                data.append(info)
            elif field_type == "region":
                info = Parser.get_field(msg, "Регион")
                data.append(info)
            elif field_type == "birthday":
                info = Parser.get_field(msg, "Дата рождения")
                data.append(info)
            elif field_type == "address":
                info = Parser.get_field(msg, "Возможные адреса", False)
                data.append(info)
            elif field_type == "possible_names":
                info = Parser.get_field(msg, "Возможные имена", False)
                data.append(info)
        return data
    @staticmethod
    def get_field(msg, field, is_flat = True):
        if msg.find(field) == -1:
            return ''
        if is_flat:
            lmsg = msg[msg.find(field)+len(field)+2:]
            return lmsg[:lmsg.find('\n')].strip()
        else:
            lmsg = msg[msg.find(field)+len(field)+3:]
            return lmsg[:lmsg.find('\n\n')].strip()
class Status(Enum):
    """Статус аккаунта"""
    STARTED = auto()
    READY = auto()
    WAITING_INFO = auto()
    WAITING_PROFILE_INFO = auto()
    COMPLETE = auto()

class Account():
    """Класс представляет собой 1 парсящий аккаунт"""
    def __init__(self, phone, is_auto_renew, manager):
        self.manager = manager
        phone = self.clean_num(phone)
        self.phone = phone
        self.is_auto_renew = is_auto_renew
        self.client = TelegramClient(f'acc_{phone}', API_ID, API_HASH)
        @self.client.on(events.NewMessage(from_users = [self.manager.BOT_NAME]))
        async def my_event_handler(event):
            await self.on_msg(event.raw_text)
        self.status = Status.STARTED
        self.attempts = 0
        self.is_sub = False
        self.numbers = []

    def get_number(self):
        if self.numbers:
            return self.numbers.pop(0)
        else:
            return False
    async def on_msg(self, msg):
        """действие при получении сообщения"""
        if self.status == Status.WAITING_PROFILE_INFO:
            try:
                self.attempts = Parser.get_attempts(msg)
                self.is_sub = True
                self.status = Status.READY
                logger.info(f"{self.phone} have {self.attempts}")
            except NotSubed:
                self.is_sub = False
                self.attempts = 0
                self.status = Status.COMPLETE
                logger.info(f"{self.phone} not subed")
            except NotExpectedMsg as e:
                logger.error(f"{self.phone} get not expected msg \n {msg}")
                return

        elif self.status == Status.WAITING_INFO:
            try:
                data = Parser.get_data(msg, self.manager.FIELDS)
                Archive.add_to_archive(data[0], msg)
                self.manager.data.append(data)
                self.manager.save_data()
                self.attempts -= 1
                logger.info(f"{self.phone} have {self.attempts} attempts")
                number = self.get_number()
                if not number:
                    self.status = Status.COMPLETE
                    logger.info(f"{self.phone} comlete parsing!")

                else:
                    await asyncio.sleep(3)
                    await self.client.send_message(self.manager. BOT_NAME, number)

            except NotExpectedMsg as e:
                # logger.error(f"{self.phone} get not expected msg while getting phone info \n {msg}")
                return

    async def login(self):
        """Логинит аккаунт"""
        logger.info(f"loggining {self.phone}")
        await self.client.start(phone = self.phone)

    async def get_info(self):
        """Запрашивает информацию об аккаунте"""
        logger.info(f"getting {self.phone} info..")
        self.status = Status.WAITING_PROFILE_INFO
        await self.client.send_message(self.manager. BOT_NAME, "👤 Мой аккаунт")

    async def start_parsing(self, numbers):
        """Начинает парсинг"""
        if not numbers:
            logger.error(f"{self.phone}| empty numbers list!")
            self.status = Status.COMPLETE
            return
        if not self.attempts:
            logger.error(f"{self.phone}| do not have any attempts!")
            self.status = Status.COMPLETE
            return
        if self.status == Status.READY:
            self.numbers = numbers
            logger.info(f"{self.phone} starting parsing {len(self.numbers)} phones")
            self.status = Status.WAITING_INFO
            await self.client.send_message(self.manager. BOT_NAME, self.get_number())
        else:
            logger.error(f"account {self.phone} not ready!")
    @staticmethod
    def clean_num(num):
        return ''.join(filter(str.isdigit, num))

class AccountManager():
    def __init__(self):
        settings = self.get_settings()
        self.BOT_NAME = settings["glaz_bot"]
        self.FIELDS = settings["fields"]
        account_infos = settings["accounts"]
        self.data = []
        self.accounts = []
        self.numbers = set()
        for info in account_infos:
            self.accounts.append(Account(info.get("phone"), info.get("is_auto_renew", False), self))

    def get_number(self):
        if self.numbers:
            return self.numbers.pop()
        else:
            return False
    
    def load_numbers(self):
        """Ищет excel файлы в папке input и берёт номера из колонки A"""
        file_path = glob.glob("./input/[!~$]*.xlsx")
        for filename in file_path:
            wb = load_workbook(filename)
            ws = wb.active
            data = ws['A']
            for item in data:
                value = item.value
                number = self.clean_ru_num(value)
                if not number:
                    logger.error(f"{value} не может быть прочитан как номер")
                else:
                    if number in self.numbers:
                        logger.info(f"{number} already loaded")
                    self.numbers.add(number)
        logger.info(f"{len(self.numbers)} numbers loaded")

    def save_data(self):
        wb = Workbook()
        ws = wb.active
        for i, values in enumerate(self.data):
            for j, val in enumerate(values):
                ws.cell(row=i+1, column=j+1, value=val)
        wb.save('result.xlsx')

    def load_from_archive(self):
        """Проверяет есть ли какие-то номера в архиве"""
        nums_in_arch = set()
        Archive.create_table()
        for number in self.numbers:
            msg = Archive.get_from_archive(number)
            if msg:
                nums_in_arch.add(number)
                self.data.append(Parser.get_data(msg, self.FIELDS))
        if self.data:
            self.save_data()
        self.numbers -= nums_in_arch
        logger.info(f"{len(nums_in_arch)} already in archive")
        logger.info(f"Need to parse: {len(self.numbers)}")

    async def init_accounts(self):
        """Авторизует и получает информацию о всех аккаунтах"""
        logger.info("logging accounts..")
        get_info_tasks = []
        for account in self.accounts:
            await account.login()
            get_info_tasks.append(account.get_info())
        logger.info("getting accounts info..")

        await asyncio.gather(*get_info_tasks)
        while True:
            is_ready = True
            for account in self.accounts:
                if account.status != Status.READY and account.status != Status.COMPLETE:
                    is_ready = False
                    logger.info(f"{account.phone} not ready")
                    await asyncio.sleep(5)
                    break
            if is_ready:
                break
        logger.info("accounts are ready to parse")

    async def start_parsing(self):
        numbers_to_parse = []
        for account in self.accounts:
            if account.attempts:
                numbers_to_parse.append({"account":account, "numbers":[]})
        numbers_to_parse.sort(key=lambda k: k['account'].attempts)
        while True:
            is_ready = True
            for acc_info in numbers_to_parse:
                number = self.get_number()
                if not number:
                    is_ready = True
                    break
                if acc_info["account"].attempts > len(acc_info["numbers"]):
                    acc_info["numbers"].append(number)
                    is_ready = False
            if is_ready:
                break

        parse_functions = []
        for acc_info in numbers_to_parse:
            parse_functions.append(acc_info["account"].start_parsing(acc_info["numbers"]))


        await asyncio.gather(*parse_functions)
        while True:
            is_ready = True
            for account in self.accounts:
                if account.status != Status.COMPLETE:
                    is_ready = False
                    await asyncio.sleep(10)
                    break
            if is_ready:
                break


    async def start(self):
        self.load_numbers()
        self.load_from_archive()
        await self.init_accounts()
        await self.start_parsing()



    @staticmethod
    def get_settings():
        with open('settings.json') as json_file:
            return json.load(json_file)

    @staticmethod
    def clean_ru_num(num):
        num = str(num)
        digits_only = ''.join(filter(str.isdigit, num))
        if len(digits_only) == 10:
            if digits_only[0] != "9":
                return False
            else:
                return "7"+digits_only
        elif len(digits_only) == 11:
            if digits_only[1] != "9":
                return False
            if digits_only[0] == "8":
                return "7"+digits_only[1:]
            elif digits_only[0] == "7":
                return digits_only
            else:
                return False
        else:
            return False


async def main():
    manager = AccountManager()
    await manager.start()


loop = asyncio.get_event_loop()
loop.run_until_complete(main())



